# %%
"""
Script that given the path to the CBTM_v2 radiology and histopatology dataset, scrapes the dataset and saved the information into a CSV file.

There are a couple of values that are important to count:
- number of radiology subjects
- number of histopathology subjects
- number of overlapping subjects

- number of sessions (pre operative and post operative)
- number of volumes for each MR modality

To accoplish this, the script needs to scrape the radiology and the histopathology datasets together.
The radiology dataset also contains a .json file for each MR acquisition. A utility can be implemented to read this and take out the needed infromation.


CBTN_v2 data size
    1.0 TiB [############] /LGG_v2
  323.9 GiB [###         ] /Medullo_v2
  178.4 GiB [##          ] /HGG_v2
  141.7 GiB [#           ] /Ependymoma_v2
   85.3 GiB [#           ] /Craniopharyngioma_v2
   76.8 GiB [            ] /ATRT_v2
   22.5 GiB [            ] /DIPG_v2


   ## Radiology

Count to be able to answer the following questions:

- [ ]  how many patients (male, female, age)
- [ ]  how many sessions (pre-op)
- [ ]  how many T1w, T1wc, T2w, FLAIR and Diffusion volumes we have per subject and per session
- [ ]  what type of diffusion data is available
- [ ]  what mean resolution does the different MR sequences have
- [ ]  from how many scanners was the data collected

## Histology

Count to be able to answer these questions:

- [ ]  how many patients (male, female, age)
- [ ]  how many sessions
- [ ]  how many stains available
- [ ]  how many of each stain available
- [ ]  how many scanners
- [ ]  which resolution each scan (and magnification)

The ultimate way to store the information is in a database that we can query. But is might take time, so for now we do it the usual excel way. Several excel tables will be created:

1. Subject level information (ID, diagnosis_folder, disgnosis_clinical_file, gender, age, survival, free_survival, nbr_sessions, has_T1w, has_T1wGD, has_T2w, was_FLAIR, has_ADC, has_HandE, has_NameOtherHistologyStains)
2. Session level radiology information (subject ID, session_ID, has_T1w, has_T1wGD, has_T2w, was_FLAIR, has_ADC)
3. File level radiology information (subject ID, session_ID, MR_sequence_type, 2D_or_3D, resolution_x, resolution_y, resolution_x, orientation, scanner)
4. Session level histology (subject ID, session_ID, has_NameOfStainings)
5. File level histopathology (subject ID, session_ID, staining_type, magnification,resolution_x, resolution_y, scanner…)

STEPS
We build the information for each subject top-down, going from the clinical files, the available radiology and histology sessions and the files.
"""

import os
import csv
import glob
import pathlib
import openpyxl
import numpy as np
import json
import pandas as pd
from copy import deepcopy


# %% UTILITIES
def get_diagnosis(primary_diagnosis, diagnosis_description):
    """
    Utility that given the list of primary diagnosis and description tries to
    identify the tumor type.

    Returns
    tumor_type : the identified tumor type. Set to None if it fails
    description : list of additional descriptions is case primaty diagnosis is Other
    grade : LGG or HGG if it can be infered by the diagnosis, else None
    """

    tumor_types = {
        "Atypical Teratoid Rhabdoid Tumor (ATRT)": "ATRT",
        "Not Reported": "None",
        "Other": "None",
        "Cavernoma": "CAVERNOMA",
        "Brainstem glioma- Diffuse intrinsic pontine glioma": "DIPG",
        "High-grade glioma/astrocytoma (WHO grade III/IV)": "ASTROCYTOMA",
        "Low-grade glioma/astrocytoma (WHO grade I/II)": "ASTROCYTOMA",
        "Germinoma": "GERMINOMA",
        "Ganglioglioma": "GANGLIOGLIOMA",
        "Teratoma": "TERATOMA",
        "Ependymoma": "EPENDYMOMA",
        "Neurofibroma/Plexiform": "NEUROFIBRILOMA-PLEXIFORM",
        "Low-grade glioma/astrocytoma (WHO grade I/II); Other": "ASTROCYTOMA-OTHERS",
        "Medulloblastoma": "MEDULLOBLASTOMA",
        "Ganglioglioma; Low-grade glioma/astrocytoma (WHO grade I/II)": "GANGLIOGLIOMA-ASTROCYTOMA",
        "Meningioma": "MENINGIOMA",
        "Embryonal Tumor with Multilayered Rosettes (ETMR); Other": "ETMR-OTHERS",
        "Embryonal Tumor with Multilayered Rosettes (ETMR)": "ETMR",
    }

    description_types = [
        "Glial neoplasm",
        "pilocytic/piloid features",
        "Rhabdoid Tumor" "Osteosarcoma" "Ependymoblastoma",
    ]

    description_types = {
        "Not Applicable": "Not-applicable",
        "Inflammation": "INFLAMATION",
        "Rhabdoid Tumor": "RHABDOID",
        "Proteinaceous debris and small fragment of adenohypophysis": "Proteinaceous debris and small fragment of adenohypophysis",
        "Lesional tissue with pilocytic/piloid features": "Lesional tissue with pilocytic/piloid features",
        "Portions of white matter with a few reactive astrocytes and rare cells with mild atypia": "Portions of white matter with a few reactive astrocytes and rare cells with mild atypia",
        "Glial neoplasm": "GLIAL-NEOPLASM",
        "Low-grade glioma": "LOW-GRADE-GIOMA",
        "Non-diagnostic tissue": "NON-DIAGNOSTIC-TISSUE",
        "Osteosarcoma": "OSTEOSARCOMA",
        "Ependymoblastoma": "EPENDYMOBLASTOMA",
    }

    # if there is no diagnosis left
    if len(primary_diagnosis) == 0:
        tumor_type = None
        description = None
        grade = "Not-available"
        return tumor_type, description, grade
    else:
        diagnosis_list = []
        description_list = []
        grade_list = []
        for idx, d in enumerate(primary_diagnosis):
            if "Other" in d:
                # check the diagnosis description
                check = [
                    diagnosis_description[idx] == t for t in description_types.keys()
                ]
                if any(check):
                    description_list.append(
                        description_types[diagnosis_description[idx]]
                    )
                    grade_list.append("Not available")
            elif not "Not Reported" in d:
                # check which of the tumor types
                check = [d == t for t in tumor_types.keys()]
                if any(check):
                    diagnosis_list.append(tumor_types[d])
                else:
                    diagnosis_list.append("Tumor type not in default")

                # check LGG or HGG
                if "Low-grade" in d:
                    grade_list.append("LGG")
                elif "High-grade" in d:
                    grade_list.append("HGG")
                else:
                    grade_list.append("Not available")

        # return a list of unique elements of diagnosis, descriptions and grades
        tumor_type = list(dict.fromkeys(diagnosis_list))
        description = list(dict.fromkeys(description_list))
        grade = list(dict.fromkeys(grade_list))

        return tumor_type, description, grade


def check_bval_bvec(volumes):
    """
    Given a list of volume that are .bvalues and .bvec, opens the files and
    checks that the information is useful (different from all ones or zeros).
    """
    # try to open .bvec file and check if information is useful
    try:
        bvec_file = [
            v
            for v in volumes
            if (".bvec" in os.path.basename(v) and ".json" not in os.path.basename(v))
        ]
        # there may be more tha one .bvec file, check all of them
        for f in bvec_file:
            aus_flag = []
            with open(f) as file:
                for line in file.readlines():
                    # print(line)
                    line = line.replace("\n", " ").replace("\t", " ")
                    aus_flag.append(
                        any(
                            [
                                float(c) != 0.0 and float(c) != 1.0
                                for c in line.strip().split(" ")
                            ]
                        )
                    )
        bvec_usability_flag = any(aus_flag)
    except:
        bvec_usability_flag = False

    # try to open .bval file and check if information is useful
    try:
        bval_file = [
            v
            for v in volumes
            if (".bval" in os.path.basename(v) and ".json" not in os.path.basename(v))
        ]
        # there may be more tha one .bvec file, check all of them
        for f in bval_file:
            aus_flag = []
            with open(f) as file:
                for line in file.readlines():
                    aus_flag.append(
                        any([c != "0" and c != "1" for c in line.strip().split(" ")])
                    )
        bval_usability_flag = any(aus_flag)
    except:
        bval_usability_flag = False

    # return the file name of the volume that is not .bval or .bvec and the usability
    diff_volume = [
        v
        for v in volumes
        if (".nii.gz" in os.path.basename(v) and ".json" not in os.path.basename(v))
    ]
    return diff_volume, any([bvec_usability_flag, bval_usability_flag])


def get_scan_type(file_path):
    """
    Heuristic that checks the file name and tries to identify the type of scan.
    Fooking for:
    - localizer: localizer
    - t1: se tw, se_t1, t1_se or t1
    - t2
    - t2 FLAIR: t2_flair
    - diffusion: diff, adc, ADC
    """

    # get only file name
    file_name = os.path.basename(file_path)
    # if 'se t1' in file_name or 'se_t1' in file_name or 't1_se' in file_name or 't1' in file_name or 'T1' in file_name:
    if any(target in file_name for target in ["se t1", "se_t1", "t1_se", "t1", "T1"]):
        scan_type = "T1"
    elif any(target in file_name for target in ["t2_flair", "T2_FLAIR"]):
        scan_type = "T2_FLAIR"
    elif any(target in file_name for target in ["t2", "T2"]):
        scan_type = "T2"
    elif any(target in file_name for target in ["flair", "FLAIR"]):
        scan_type = "FLAIR"
    elif any(
        target in file_name
        for target in ["diff", "DIFF", "trace", "TRACE", "adc", "ADC", "EXP", "FA"]
    ):
        scan_type = "DIFFUSION"
        # there are different types of diffusion files
        if any(target in file_name for target in ["TRACE"]):
            scan_type = "DIFFUSION_TRACE"
        elif any(target in file_name for target in ["adc", "ADC"]):
            scan_type = "DIFFUSION_ADC"
        elif any(target in file_name for target in ["EXP"]):
            scan_type = "DIFFUSION_EXP"
        elif any(target in file_name for target in ["FA"]):
            scan_type = "DIFFUSION_FA"
    elif any(target in file_name for target in ["localizer"]):
        scan_type = "LOCALIZER"
    elif any(target in file_name for target in ["Perfusion"]):
        scan_type = "PERFUSION"
    else:
        scan_type = "UNKNOWN"

    return scan_type


def get_file_extension(files):
    """
    Utility that given a list of files returs a list of extensions of the files
    in the input file list
    """

    extensions = []
    for f in files:
        aus = pathlib.Path(f).suffixes
        if len(aus) == 1:
            extensions.append(aus[0])
        elif len(aus) >= 2:
            if aus[-1] == ".json":
                extensions.append(aus[-1])
            else:
                extensions.append("".join(aus[-2::]))

    return extensions


def check_pre_or_post_operative_session(
    clinical_timeline, clinical_diagnosis_type, session_time
):
    """
    Utility function that uses the information in the clinical information file
    to understand if the session is pre or post operative. In particular, using
    the diagnostic_type and the age of the patient (in days) at which it was
    done, one can see if it is the first encounter of the tumor, recurrence
    (defined as tumor re-growth after total resection) and progressive (defined
    as tumor growth after partial resection). Thus, scans that are intermediate
    in time between the initial enounter and recurrence/progressive, are pre
    operative.

    It is important also to check that, if there is only one the status will be
    unknown if the session_time is antecedent the first encounter. If the
    session time is precedent the first encounter, the session status is set to
    pre_op.

    Steps
    - find the time of the Initial CNS Tumor (first encounter)
    - find the time of the first Recurrence or Progression
    - check where the session time locates compared to the initial encounter
        and the recurrence or progression
    """
    try:
        first_encounter_time_idx = max(
            loc
            for loc, val in enumerate(clinical_diagnosis_type)
            if val == "Initial CNS Tumor"
        )
    except:
        # if there is no record of the first encounter
        return "unknown"

    """ from previous version
    if (
        len(clinical_diagnosis_type) > 1
        and first_encounter_time_idx < len(clinical_diagnosis_type) - 1
    ):
        # there are multiple clinical_diagnosis_type and the first encounter is not the last encounter
        if session_time < clinical_timeline[first_encounter_time_idx + 1]:
            return "pre_op"
        else:
            return "post_op"
    elif session_time <= clinical_timeline[first_encounter_time_idx]:
        return "pre_op"
    else:
        return "post_op"
    """
    if session_time <= clinical_timeline[first_encounter_time_idx]:
        return "pre_op"
    else:
        return "post_op"


def get_contrast_status(file_path):
    """
    Utility that given a file path, tries to find out if the scan was performed
    pre or post contrast agent. Here is mostly important for T1 scans, since
    all the other acquisitions should be pre contrast
    """
    # get only file name
    file_name = os.path.basename(file_path)
    # if 'se t1' in file_name or 'se_t1' in file_name or 't1_se' in file_name or 't1' in file_name or 'T1' in file_name:
    if any(target in file_name for target in ["post", "POST", "P0ST", "Post"]):
        contrast_status = "post_contrast"
    else:
        contrast_status = "pre_contrast"

    return contrast_status


def get_empty_overall_subject_info():
    return {
        "sex": "Not_available",
        "age_at_diagnosis": "Not_available",
        "diagnosis": "Not_available",
        "tumor_grade": "Not_available",
        "overall_survival": "Not_available",
        "progression_free_survival": "Not_available",
        "vital_status": "Not_available",
        "age_at_sample_acquisition": [],
        "tumor_descriptor": [],
    }


def get_empty_radiology_session_dict():
    return {
        "T1w": [],
        "T1wGD": [],
        "T2w": [],
        "FLAIR": [],
        "DIFFUSION": [],
        "FSPGR": [],
        "UNKNOWN": [],
        "PERFUSION": [],
        "SWI": [],
        "MAG": [],
        "PHE": [],
        "ASL": [],
        "SWAN": [],
        "CBF": [],
        "pre_post_operation_status": None,
    }


def get_empty_histology_session_dict():
    return {
        "HandE": [],
        "EMA": [],  # Epithelial membrane antigen
        "GFAP": [],  # Immunohistochemical staining for glial fibrillary acidic protein
        "KI-67": [],
        "RETICULIN": [],  # staining that stains the reticulinic fibers of the stroma and makes it possible to specify the architecture of the tumors
        "VIMENTIN": [],  # stains virtually all spindle cell neoplasms—mesenchymal spindle cell neoplasms and sarcomatoid carcinomas included
        "INI-1": [],  # nuclear staining
        "BRG1": [],  # Immunohistochemical analysis of SMARCA4 / BRG1
        "SMA": [],  # Smooth Muscle Actin
        "SYNAPTO": [],  # Synaptophysin is an integral membrane protein localized to synaptic vesicles
        "NFP": [],  # neurofilament
        "OLIG2": [],  # Anti-Human Olig2 Antibody
        "TTR": [],  # Transthyretin
        "UNKNOWN": [],
    }


def get_radiology_acquisition_information(path_to_acquisition_json_file):
    # build empty acquisition disctionary
    acquisition_info_dict = {
        "2D_or_3D": None,
        "orientation": None,
        "resolution_dim1": None,
        "resolution_dim2": None,
        "resolution_dim3": None,
        "scanner": None,
    }
    try:
        with open(path_to_acquisition_json_file) as json_file:
            acquisition_information = json.load(json_file)
            available_json_file_flag = True

            # polulate the values
        for key_1, key_2 in zip(
            [
                "2D_or_3D",
                "orientation",
                "resolution_dim1",
                "resolution_dim2",
                "resolution_dim3",
                "scanner",
            ],
            [
                "MRAcquisitionType",
                "PhaseEncodingDirection",
                "dim1",
                "dim2",
                "dim3",
                ["Manufacturer", "ManufacturersModelName"],
            ],
        ):
            try:
                if key_1 != "scanner":
                    acquisition_info_dict[key_1] = acquisition_information[key_2]
                elif all([key_1 == "2D_or_3D", acquisition_information[key_2] == "3D"]):
                    acquisition_info_dict["resolution_dim1"] = acquisition_information[
                        "SliceThickness"
                    ]
                    acquisition_info_dict["resolution_dim2"] = acquisition_information[
                        "SliceThickness"
                    ]
                    acquisition_info_dict["resolution_dim3"] = acquisition_information[
                        "SliceThickness"
                    ]
                else:
                    acquisition_info_dict[key_1] = (
                        acquisition_information[key_2[0]]
                        + "_"
                        + acquisition_information[key_2[1]]
                    )
            except:
                acquisition_info_dict[key_1] = "Not_available_in_file"

    except:
        available_json_file_flag = False

    # add information about file name
    acquisition_info_dict["file_name"] = os.path.join(
        path_to_acquisition_json_file.split(os.path.sep)[-3],
        path_to_acquisition_json_file.split(os.path.sep)[-2],
        path_to_acquisition_json_file.split(os.path.sep)[-1],
    )

    # get the MR sequence type (use name and .json file information)
    if available_json_file_flag:
        acqusition_name = acquisition_information["SeriesDescription"]
    else:
        # infere from the file name
        acqusition_name = path_to_acquisition_json_file.split(os.path.sep)[-1].split(
            "."
        )[0]

    def get_acquisition_type(acqusition_name):
        """
        Heuristic that checks the file name and tries to identify the type of scan.
        Fooking for:
        - localizer: localizer
        - t1: se tw, se_t1, t1_se or t1
        - t2
        - t2 FLAIR: t2_flair
        - diffusion: diff, adc, ADC
        - FSPGR (fast spoiled gradient-echo dual echo): FSPGR
        - SWI (Susceptibility weighted imaging): SWI
        - Mag (magnitude image): Mag
        - Phe (phase image): Phe
        - ASL (arterial spin labeling): ASL
        - SWAN (susceptibility-weighted angiography):SWAN
        - CBF (cerebral blood flow): CBF
        """

        # if 'se t1' in file_name or 'se_t1' in file_name or 't1_se' in file_name or 't1' in file_name or 'T1' in file_name:
        if any(
            target in acqusition_name
            for target in ["se t1", "se_t1", "t1_se", "t1", "T1"]
        ):
            scan_type = "T1w"
        elif any(
            target in acqusition_name
            for target in ["t2_flair", "T2_FLAIR", "flair", "FLAIR"]
        ):
            scan_type = "FLAIR"
        elif any(target in acqusition_name for target in ["t2", "T2"]):
            scan_type = "T2w"
        elif any(target in acqusition_name for target in ["FSPGR", "fspgr"]):
            scan_type = "FSPGR"
        elif any(target in acqusition_name for target in ["SWI", "swi"]):
            scan_type = "SWI"
        elif any(target in acqusition_name for target in ["Mag", "MAG", "mag"]):
            scan_type = "MAG"
        elif any(target in acqusition_name for target in ["PHE", "Phe", "phe"]):
            scan_type = "PHE"
        elif any(target in acqusition_name for target in ["ASL", "asl"]):
            scan_type = "ASL"
        elif any(target in acqusition_name for target in ["SWAN", "swan"]):
            scan_type = "SWAN"
        elif any(
            target in acqusition_name
            for target in ["CBF", "cbf", "Cerebral_Blood_Flow"]
        ):
            scan_type = "CBF"
        elif any(
            target in acqusition_name
            for target in [
                "diff",
                "DIFF",
                "trace",
                "TRACE",
                "adc",
                "ADC",
                "EXP",
                "FA",
                "Apparent_Diffusion_Coefficient",
                "DWI",
            ]
        ):
            scan_type = ("DIFFUSION", "")
            # there are different types of diffusion files
            if any(target in acqusition_name for target in ["TRACE", "trace"]):
                scan_type = ("DIFFUSION", "TRACE")
            elif any(
                target in acqusition_name
                for target in ["adc", "ADC", "Apparent_Diffusion_Coefficient"]
            ):
                scan_type = ("DIFFUSION", "ADC")
            elif any(target in acqusition_name for target in ["EXP"]):
                scan_type = ("DIFFUSION", "EXP")
            elif any(target in acqusition_name for target in ["FA"]):
                scan_type = ("DIFFUSION", "FA")
        elif any(target in acqusition_name for target in ["localizer"]):
            scan_type = "LOCALIZER"
        elif any(target in acqusition_name for target in ["Perfusion"]):
            scan_type = "PERFUSION"
        else:
            scan_type = "UNKNOWN"

        return scan_type

    def get_contrast_status(acqusition_name):
        """
        Utility that given a file path, tries to find out if the scan was performed
        pre or post contrast agent. Here is mostly important for T1 scans, since
        all the other acquisitions should be pre contrast
        """

        # if 'se t1' in file_name or 'se_t1' in file_name or 't1_se' in file_name or 't1' in file_name or 'T1' in file_name:
        if any(
            target in acqusition_name
            for target in ["post", "POST", "P0ST", "Post", "C+"]
        ):
            contrast_status = "post_contrast"
        else:
            contrast_status = "pre_contrast"

        return contrast_status

    MR_sequence = get_acquisition_type(acqusition_name)

    # if DIFFUSION add infromation about the type of diffusion
    if isinstance(MR_sequence, tuple):
        acquisition_info_dict["diffusion_type"] = MR_sequence[1]
        MR_sequence = MR_sequence[0]

    # check if pre or post contrast  (change T1w to T1wGD if contrast is used)
    contrast_status = get_contrast_status(acqusition_name)
    acquisition_info_dict["GD_contrast"] = contrast_status

    if all([contrast_status == "post_contrast", MR_sequence == "T1w"]):
        MR_sequence = "T1wGD"

    # return values
    return MR_sequence, acquisition_info_dict


def get_histology_information(json_information_file):
    # build empty acquisition disctionary
    acquisition_info_dict = {
        "magnification": None,  # AppMag
        "resolution_dim1": None,  # mpp-x
        "resolution_dim2": None,  # mpp-y
        "scanner": None,  # vendor
        "level_count": None,  # level-count
    }
    try:
        with open(json_information_file) as json_file:
            acquisition_information = json.load(json_file)
            available_json_file_flag = True

            # polulate the values
        for key_1, key_2 in zip(
            [
                "magnification",
                "resolution_dim1",
                "resolution_dim2",
                "scanner",
                "level_count",
            ],
            [
                "AppMag",
                "mpp-x",
                "mpp-y",
                "vendor",
                "level-count",
            ],
        ):
            try:
                acquisition_info_dict[key_1] = acquisition_information["info"][key_2]
            except:
                acquisition_info_dict[key_1] = "Not_available_in_file"
    except:
        available_json_file_flag = False

    # add information about file name
    acquisition_info_dict["file_name"] = json_information_file.split(os.path.sep)[-1]

    # get file name and use to identify the stain type
    file_name = os.path.basename(json_information_file).split(".")[0]

    def get_stain_type(file_name):
        """
        Heuristic to get the stain type
        Fooking for:
        - HandE
        - EMA: Epithelial membrane antigen
        - GFAP: Immunohistochemical staining for glial fibrillary acidic protein
        - KI-67
        - RETICULIN: staining that stains the reticulinic fibers of the stroma and makes it possible to specify the architecture of the tumors
        - VIMENTIN: stains virtually all spindle cell neoplasms—mesenchymal spindle cell neoplasms and sarcomatoid carcinomas included
        - INI-1: nuclear staining
        - BRG1: Immunohistochemical analysis of SMARCA4 / BRG1
        - SMA: Smooth Muscle Actin
        - SYNAPTO: Synaptophysin is an integral membrane protein localized to synaptic vesicles
        - NFP: neurofilament
        - OLIG2: Anti-Human Olig2 Antibody
        - UNKNOWN
        """

        stains = {
            "HandE": ["HandE", "hande", "HANDE", "H_and_E"],
            "EMA": ["EMA", "ema"],
            "GFAP": ["GFAP", "gfap"],
            "KI-67": ["KI-67", "ki-67", "Ki-67", "KI67"],
            "INI-1": ["INI-1", "ini-1"],
            "SMA": ["SMA", "sma", "Sma"],
            "SYNAPTO": ["Synapto", "SYNAPTO"],
            "NFP": ["NFP", "nfp"],
            "BRG1": ["BRG-1", "BRG1", "brg-1", "brg1"],
            "RETICULIN": ["RETICULIN", "reticulin"],
            "VIMENTIN": ["VIMENTIN", "vimentin", "Vimentin", "Vim_", "vim_", "VIM"],
            "OLIG2": ["OLIG-2", "Olig-2"],
            "TTR": ["TTR"],
        }

        histo_stain = "UNKNOWN"
        for stain_name, possible_stain_names in stains.items():
            if any(
                [
                    possible_stain_name in file_name
                    for possible_stain_name in possible_stain_names
                ]
            ):
                histo_stain = stain_name
        return histo_stain

    # get stain type
    stain_type = get_stain_type(file_name)

    return stain_type, acquisition_info_dict


# %% SOME INFORMATION
""" INFORMATION ABOUT DATA STRUCTURE
to get all this information we can do many passes through the dataset and collect the information that
is needed, or collect all the infromation in one go, and fill in the infromation in the different files by
going through the saved information.

Here the second option is used, where all the information about from one subject is collected in a dictionary
subject_ID
    gender
    age_at_diagnosys
    diagnosis_from_folder
    diagnosis_from_clinical_file
    tumor_grade
    overall_survival
    progressing_free_survival
    radiology_sessions
        session_ID
            T1w_sequences
             [
                {"2D_or_3D",
                    "orientation",
                    "resolution_dim1",
                    "resolution_dim2",
                    "resolution_dim3",
                    "scanner",
                }
                {"2D_or_3D",
                    "orientation",
                    "resolution_dim1",
                    "resolution_dim2",
                    "resolution_dim3",
                    "scanner",
                }
             ]
            T1wGD_sequences
             [
                {},
                {},
             ]
            T2w_sequences
            FLAIR_sequences
            Diffusion_sequences
    histology_sessions
        session_ID
            HandE_staining
             [
                {    
                    "magnification",
                    "resolution_dim1",
                    "resolution_dim2",
                    "scanner",
                },
             ]
            Ki67
             [
                {},
                {},
             ]
            OtherStainings
                {},
                {},
"""

# %% GET UNIQUE SUBJECT IDs FROM THE RADIOLOGY AND HISTOPATHOLOGY DATASET
radiology_dataset_path = "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/RADIOLOGY"
histology_dataset_path = "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/HISTOLOGY"

tumor_types = {
    "ATRT": "ATRT_v2",
    "DIPG": "DIPG_v2",
    "HGG": "HGG_v2",
    "MEDULLOBLASTOMA": "Medullo_v2",
    "CRANIOPHARYNGIOMA": "Craniopharyngioma_v2",
    "EPENDYMOMA": "Ependymoma_v2",
    "LGG": "LGG_v2",
}

# each element in the subject ID has the ID, tumor_type based on the radiology folder, radiology_folder, histology_folder)
subject_IDs = {}

# through the radiology dataset
for tumor_type, folder_name in tumor_types.items():
    subjects = [
        s.split(os.path.sep)[-2]
        for s in glob.glob(
            os.path.join(radiology_dataset_path, folder_name, "SUBJECTS", "*/")
        )
    ]
    # add to the dictionary
    for s in list(dict.fromkeys(subjects)):
        subject_IDs[s] = {
            "tumor_type": tumor_type,
            "radiology_folder": os.path.join(
                radiology_dataset_path, folder_name, "SUBJECTS"
            ),
            "histology_folder": None,
        }
    print(f"Tumor Type: {tumor_type} -> {len(subjects)} subjects")
# get the unique values

# through the histology dataset
subjects = [
    s.split(os.path.sep)[-2]
    for s in glob.glob(os.path.join(histology_dataset_path, "SUBJECTS", "*/"))
]
# add to the dictionary
for s in list(dict.fromkeys(subjects)):
    try:
        subject_IDs[s]["histology_folder"] = os.path.join(
            histology_dataset_path, "SUBJECTS"
        )
    except:
        subject_IDs[s] = {
            "tumor_type": None,
            "radiology_folder": None,
            "histology_folder": os.path.join(histology_dataset_path, "SUBJECTS"),
        }

# %% OPEN CLINICAL FILES (both the ones for each tumor and the ones from the portal)
"""
The files from the portal should cover all the subject IDs available in the radiology and histology folders.
Thus we use these as starting point to gather the information from every subject. The information from the portal
files are integrated with the ones specific for each tumor (if the tumor type and subject is available). 
"""
# define paths of clinical files (change based on your location)
clinical_file_from_portal_path = "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/CSV_FILES/CBTN_clinical_data_from_portal.xlsx"
biospecimen_file_from_portal_path = "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/CSV_FILES/CBNT_biospecimenData_from_portal.xlsx"

clinical_info_files = {
    "ATRT": "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/CSV_FILES/ATRT_clinical_information.xlsx",
    "DIPG": "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/CSV_FILES/DIPG_clinical_information.xlsx",
    "LGG": "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/CSV_FILES/LGG_clinical_information.xlsx",
    "MEDULLO": "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/CSV_FILES/Medullo_clinical_information.xlsx",
}

# open the files (just open)
print("Opening clinical and biospecimens file from Kids First portal...")
clinical_file_from_portal_patients_info = pd.read_excel(
    clinical_file_from_portal_path, sheet_name="Participants", index_col="External Id"
)
clinical_file_from_portal_diagnosis_info = pd.read_excel(
    clinical_file_from_portal_path, sheet_name="Diagnoses", index_col="External Id"
)
biospecimen_file_from_portal = pd.read_excel(
    biospecimen_file_from_portal_path,
    sheet_name="Biospecimens",
    index_col="External Id",
)

# filter to only use CBTNs data
clinical_file_from_portal_patients_info = clinical_file_from_portal_patients_info[
    clinical_file_from_portal_patients_info["Study"]
    == "Pediatric Brain Tumor Atlas: CBTTC"
]
clinical_file_from_portal_diagnosis_info = clinical_file_from_portal_diagnosis_info[
    clinical_file_from_portal_diagnosis_info.index.isin(
        clinical_file_from_portal_patients_info.index
    )
]
biospecimen_file_from_portal = biospecimen_file_from_portal[
    biospecimen_file_from_portal.index.isin(
        clinical_file_from_portal_patients_info.index
    )
]
print("Done!")

print("Opening per-tumor type clinical files...")
clinical_file_per_tumor = pd.concat(
    [
        pd.read_excel(f, sheet_name=1, index_col=0)
        for f in clinical_info_files.values()
        if os.path.isfile(f)
    ]
)
print("Done!")


# %% SCRAPE CLINICAL FILES (only the subjects that appear in the radiology and histology folders)

# for each subject we save the different information
all_information = {}

"""
Open the clinical_file_from_portal and get the unique subjects.
For the unique subjects, extract the information that is available from the clinical_file_from_portal (sex, diagnosis, age_at_diagnosis, overall survival)
    Also get the information regarding the age_at_sample_acquisition alligned with the tumor_descriptor from the biospecimen_file_from_portal
    Get progression_free_survival information (if avvailable) from the detailed clinical files

NB! 
There are subject_IDs which have multiple versions in the Participants list. However, only one version of them has
the information. Thus, one needs to drop all those cases to avoid errors in the code later
"""

aus_len = len(str(len(clinical_file_from_portal_patients_info.index.unique())))
for idx, subject in enumerate(clinical_file_from_portal_patients_info.index.unique()):
    print(
        f"Working on subject {idx+1:0{aus_len}d}/{len(clinical_file_from_portal_patients_info.index.unique())} \r",
        end="",
    )
    if subject in subject_IDs.keys():
        # build entry in the all_information dictionary
        all_information[subject] = get_empty_overall_subject_info()

        # get information from the clinical_file_from_portal_patients_info
        for ikey, ykey in zip(
            ["sex", "overall_survival", "vital_status"],
            ["Gender", "Age at the Last Vital Status (Days)", "Vital Status"],
        ):
            try:
                all_information[subject][
                    ikey
                ] = clinical_file_from_portal_patients_info.loc[subject][ykey]
            except:
                continue

        # get information from the clinical_file_from_portal_diagnosis_info
        try:
            all_information[subject]["age_at_diagnosis"] = np.min(
                clinical_file_from_portal_diagnosis_info.loc[[subject]][
                    "Age at Diagnosis (Days)"
                ]
            )
        except:
            continue

        try:
            all_information[subject][
                "diagnosis"
            ] = clinical_file_from_portal_diagnosis_info.loc[[subject]][
                "Diagnosis (Source Text)"
            ].drop_duplicates()[
                0
            ]
        except:
            continue

        # get information from the biospecimen_file_from_portal
        aus_info = (
            biospecimen_file_from_portal.loc[[subject]][
                ["Age at Sample Acquisition", "Tumor Descriptor"]
            ]
            .dropna(subset=["Age at Sample Acquisition"])
            .sort_values(by=["Age at Sample Acquisition"])
            .drop_duplicates(subset=["Age at Sample Acquisition", "Tumor Descriptor"])
        )
        all_information[subject]["age_at_sample_acquisition"] = list(
            aus_info["Age at Sample Acquisition"]
        )

        all_information[subject]["tumor_descriptor"] = list(
            aus_info["Tumor Descriptor"]
        )

        # from the per-tumor type clinical files get the progression_free_survival and
        # check that the survival age is smaller than the one olready availabel. If not subsitute.
        if subject in clinical_file_per_tumor.index:
            all_information[subject][
                "progression_free_survival"
            ] = clinical_file_per_tumor.loc[[subject]]["Progression Free Survival"]

            # get overall survival (these can be many values as well as 'Not Reported')
            overall_survival_from_per_tumor_clinical_file = clinical_file_per_tumor.loc[
                subject
            ]["Overall Survival"]
            if isinstance(overall_survival_from_per_tumor_clinical_file, str):
                overall_survival_from_per_tumor_clinical_file = 0
            else:
                overall_survival_from_per_tumor_clinical_file = np.max(
                    overall_survival_from_per_tumor_clinical_file
                )
            # check overall survival
            if (
                overall_survival_from_per_tumor_clinical_file
                > all_information[subject]["overall_survival"]
            ):
                all_information[subject][
                    "overall_survival"
                ] = clinical_file_per_tumor.loc[subject]["Overall Survival"]

        # if idx == 100:
        #     break

# and finally check that all the files that have been included
missing_info = []
for subject_ID in subject_IDs.keys():
    if subject_ID not in all_information.keys():
        missing_info.append(subject_ID)
if len(missing_info) != 0:
    print(
        f"Missing clinical information for {len(missing_info)} subjects. Adding information..."
    )
    for subject in missing_info:
        all_information[subject] = get_empty_overall_subject_info()


# %% LOOP THROUGH THE UNIQUE Subjects AND SCRAPE THE DATA (from the dataset folders)

for subj_idx, (subject, subj_path_info) in enumerate(subject_IDs.items()):
    # create entry in the all_information and add the subject-related information
    tumor_type_from_folder, radiology_folder, histology_folder = (
        subj_path_info["tumor_type"],
        subj_path_info["radiology_folder"],
        subj_path_info["histology_folder"],
    )

    all_information[subject]["diagnosis_from_folder"] = (
        tumor_type_from_folder if tumor_type_from_folder else "None"
    )

    # now work on scraping the radiology information if available
    all_information[subject]["nbr_radiology_sessions"] = {
        "pre_op": 0,
        "post_op": 0,
        "unknown": 0,
    }
    if radiology_folder:
        session_iter = glob.glob(
            os.path.join(radiology_folder, subject, "SESSIONS", "*/")
        )
        if len(session_iter) != 0:
            # add space for the radiology sessions
            all_information[subject]["radiology_sessions"] = {}
            # work on every session
            for session_idx, session in enumerate(session_iter):
                # get the session name from the path
                session_name = session.split(os.path.sep)[-2]
                # only process those that have brain in the name
                if not "B_brain" in session_name:
                    continue
                else:
                    # create entry in the all_information dict for this session
                    all_information[subject]["radiology_sessions"][
                        session_name
                    ] = get_empty_radiology_session_dict()

                    # get if the session is pre or post operative (this works for the subjects that have clinical information)
                    try:
                        session_status = check_pre_or_post_operative_session(
                            all_information[subject]["age_at_sample_acquisition"],
                            all_information[subject]["tumor_descriptor"],
                            int(session_name.split("d")[0]),
                        )
                        all_information[subject]["radiology_sessions"][session_name][
                            "pre_post_operation_status"
                        ] = session_status
                        # count session
                        all_information[subject]["nbr_radiology_sessions"][
                            session_status
                        ] += 1
                    except KeyError:
                        all_information[subject]["radiology_sessions"][session_name][
                            "pre_post_operation_status"
                        ] = "unknown"
                        all_information[subject]["nbr_radiology_sessions"][
                            "unknown"
                        ] += 1
                    # work on every acquisition in the session
                    acquisition_iter = glob.glob(
                        os.path.join(session, "ACQUISITIONS", "*/")
                    )
                    for acquisition_idx, acquisition in enumerate(acquisition_iter):
                        # get acquisition name
                        acquisition_name = acquisition.split(os.path.sep)[-2]
                        # do not scrape the acqusitions which are localize
                        if "localizer" in acquisition_name:
                            continue
                        else:
                            # open the .json file with all the information about the acquisition
                            # here the file that we want to open is the one that has the same name as the .nii.gz file but with .json extention
                            # get the files with .nii.gz extention
                            niigz_files = glob.glob(
                                os.path.join(acquisition, "FILES", "*.nii.gz")
                            )
                            # there can be multiple .nii.gz files, thus process each independetly
                            for niigz_file_idx, niigz_file in enumerate(niigz_files):
                                # get file name
                                file_name = niigz_file.split(os.path.sep)[-1]
                                # get information from the acquisition .json file
                                acquisition_information_file = os.path.join(
                                    acquisition,
                                    "FILES",
                                    file_name.split(".nii.gz")[0] + ".json",
                                )
                                (
                                    MR_sequence,
                                    acquisition_info,
                                ) = get_radiology_acquisition_information(
                                    acquisition_information_file
                                )
                                # trim the MR_sequence information if DIFFUSION
                                if isinstance(MR_sequence, tuple):
                                    MR_sequence = MR_sequence[0]
                                # save the infromation in the all_information
                                all_information[subject]["radiology_sessions"][
                                    session_name
                                ][MR_sequence].append(acquisition_info)

    # now work on getting the histology information if available
    all_information[subject]["nbr_histology_sessions"] = 0
    if histology_folder:
        session_iter = glob.glob(
            os.path.join(histology_folder, subject, "SESSIONS", "*/")
        )
        all_information[subject]["nbr_histology_sessions"] = len(session_iter)
        if len(session_iter) != 0:
            # add space for the histology sessions
            all_information[subject]["histology_sessions"] = {}
            # work on every session
            for session_idx, session in enumerate(session_iter):
                # get the session name from the path
                session_name = session.split(os.path.sep)[-2]
                # create entry in the all_information dict
                all_information[subject]["histology_sessions"][
                    session_name
                ] = get_empty_histology_session_dict()

                # work on every histo file in the session
                files_iter = glob.glob(
                    os.path.join(session, "ACQUISITIONS", "Files", "FILES", "*.svs")
                )
                for file_idx, file in enumerate(files_iter):
                    # get acquisition name
                    file_name = file.split(os.path.sep)[-1]
                    # open the .json file and get information about this histo image
                    information_file = os.path.join(
                        os.path.dirname(file), file_name + ".flywheel.json"
                    )
                    stain_type, histo_information = get_histology_information(
                        information_file
                    )

                    # add stain name in the disctionary if not present
                    if not any(
                        [
                            stain_type
                            == all_information[subject]["histology_sessions"][
                                session_name
                            ].keys()
                        ]
                    ):
                        all_information[subject]["histology_sessions"][session_name][
                            stain_type
                        ] = []

                    # save the infromation in the all_information
                    all_information[subject]["histology_sessions"][session_name][
                        stain_type
                    ].append(histo_information)

    # print status
    print(
        f"Processing subject {subj_idx:04d}/{len(subject_IDs)}\r",
        end="",
    )

    # if subj_idx == 100:
    #     break

# %% save all information to file

for s, i in all_information.items():
    for ss, j in i.items():
        if isinstance(j, pd.Series):
            all_information[s][ss] = list(j)

with open(
    "/run/media/chrsp39/CBNT_v2/Datasets/CBTN_v2/SUMMARY_FILES/summary_scraped_dataset.json",
    "w",
) as fp:
    json.dump(all_information, fp, sort_keys=True)

# %% BUILD THE SUBJECT LEVEL CSV FILE

for s, i in all_information.items():
    for ss, j in i.items():
        if isinstance(j, pd.Series):
            all_information[s][ss] = list(j)

csv_header = [
    "subject_ID",
    "diagnosis",
    "diagnosis_from_folder",
    "tumor_grade",
    "sex",
    "age_at_diagnosis",
    "survival",
    "free_survival",
    "session_status",
    "nbr_radiology_sessions",
    "nbr_T1w",
    "nbr_T1wGD",
    "nbr_T2w",
    "nbr_FLAIR",
    "nbr_DIFFUSION",
    "nbr_TRACE",
    "nbr_ADC",
    "nbr_FA",
    "nbr_EXP",
    "nbr_FSPGR",
    "nbr_UNKNOWN",
    "nbr_PERFUSION",
    "nbr_SWI",
    "nbr_MAG",
    "nbr_PHE",
    "nbr_ASL",
    "nbr_SWAN",
    "nbr_CBF",
    "nbr_histology_sessions",
    "nbr_HandE",
    "nbr_EMA",
    "nbr_GFAP",
    "nbr_KI-67",
    "nbr_RETICULIN",
    "nbr_VIMENTIN",
    "nbr_INI-1",
    "nbr_BRG1",
    "nbr_SMA",
    "nbr_SYNAPTO",
    "nbr_NFP",
    "nbr_OLIG2",
    "nbr_TTR",
    "nbr_UNKNOWN",
]
csv_rows = []

# some lists to make thigs slimmer
MR_sequences = [
    "T1w",
    "T1wGD",
    "T2w",
    "FLAIR",
    "DIFFUSION",
    "FSPGR",
    "PERFUSION",
    "SWI",
    "MAG",
    "PHE",
    "ASL",
    "SWAN",
    "CBF",
]
DIFFUSION_maps = ["TRACE", "ADC", "EXP", "FA"]
HISTOLOGY_stainings = [
    "HandE",
    "EMA",
    "GFAP",
    "KI-67",
    "RETICULIN",
    "VIMENTIN",
    "INI-1",
    "BRG1",
    "SMA",
    "SYNAPTO",
    "NFP",
    "OLIG2",
    "TTR",
    "UNKNOWN",
]


def get_empty_row_dict(csv_header):
    aus = {}
    for key in csv_header:
        aus[key] = None

    return aus


# loop throught the information
for idx, (subject_ID, subject_information) in enumerate(all_information.items()):
    # aus variable for saving
    saved_hist_only_row = False

    # one row for every session_status (histology information might be repeated)
    for session_status in ["pre_op", "post_op", "unknown"]:
        # initialize empty row
        aus_row_dict = dict.fromkeys(csv_header, None)

        # gather clinical_level information
        aus_row_dict["subject_ID"] = subject_ID
        aus_row_dict["diagnosis"] = subject_information["diagnosis"]
        aus_row_dict["diagnosis_from_folder"] = subject_information[
            "diagnosis_from_folder"
        ]
        aus_row_dict["tumor_grade"] = subject_information["tumor_grade"]
        aus_row_dict["age_at_diagnosis"] = subject_information["age_at_diagnosis"]
        aus_row_dict["sex"] = subject_information["sex"]
        aus_row_dict["survival"] = subject_information["overall_survival"]
        aus_row_dict["free_survival"] = subject_information["progression_free_survival"]

        # gather histology level information (always)
        aus_row_dict["nbr_histology_sessions"] = subject_information[
            "nbr_histology_sessions"
        ]

        # here count across the different sessions how many of the different histology stainings
        if subject_information["nbr_histology_sessions"] != 0:
            for hist_session in subject_information["histology_sessions"].items():
                for stain in HISTOLOGY_stainings:
                    aus_row_dict["nbr_" + stain] = (
                        (aus_row_dict["nbr_" + stain] + len(hist_session[1][stain]))
                        if aus_row_dict["nbr_" + stain]
                        else len(hist_session[1][stain])
                    )

        # here count across the different radiology sessions
        aus_row_dict["nbr_radiology_sessions"] = subject_information[
            "nbr_radiology_sessions"
        ][session_status]

        # here count across the different sessions how many of the different MR sequences
        if aus_row_dict["nbr_radiology_sessions"] != 0:
            aus_row_dict["session_status"] = session_status
            for rad_session in subject_information["radiology_sessions"].values():
                if rad_session["pre_post_operation_status"] == session_status:
                    for sequence in MR_sequences:
                        aus_row_dict["nbr_" + sequence] = (
                            (
                                aus_row_dict["nbr_" + sequence]
                                + len(rad_session[sequence])
                            )
                            if aus_row_dict["nbr_" + sequence]
                            else len(rad_session[sequence])
                        )
                        # handle the diffusion case
                        if sequence == "DIFFUSION":
                            # count the different diffusion sequences
                            for dm in DIFFUSION_maps:
                                # get the length of the diffusion sequences of this diffusion type
                                len_diff_type = len(
                                    [
                                        1
                                        for ds in rad_session[sequence]
                                        if ds["diffusion_type"] == dm
                                    ]
                                )
                                # save count
                                aus_row_dict["nbr_" + dm] = (
                                    (aus_row_dict["nbr_" + dm] + len_diff_type)
                                    if aus_row_dict["nbr_" + dm]
                                    else len_diff_type
                                )
        # Save information. Here we save a new row for every subject and session status (if there are sessions in this status)
        # The histology information is repeated since it is always post operative.
        # Save only once if only histology information
        if any(
            [
                aus_row_dict["nbr_histology_sessions"] != 0,
                aus_row_dict["nbr_radiology_sessions"] != 0,
            ]
        ):
            # check if it is an ony histology row:
            if aus_row_dict["nbr_radiology_sessions"] == 0:
                if not saved_hist_only_row:
                    csv_rows.append([aus_row_dict[k] for k in csv_header])
                    saved_hist_only_row = True
            else:
                csv_rows.append([aus_row_dict[k] for k in csv_header])

# save information
with open(
    f"/run/media/iulta54/CBNT_v2/Datasets/CBTN_v2/SUMMARY_FILES/CBTN_v2_subject_level_summary.csv",
    "w",
) as csv_file:
    wr = csv.writer(csv_file, quoting=csv.QUOTE_ALL)
    wr.writerow(csv_header)
    wr.writerows(csv_rows)


# %% Load information from the file (if needed)

with open(
    "/run/media/iulta54/GROUP_lnx1/Datasets/CBTN_v2/summary_scraped_dataset.json", "r"
) as fp:
    all_information = json.load(fp)


# %%
for tumor_type, file_path in clinical_info_files.items():
    # check if file is available, if yes open and read information
    if os.path.isfile(file_path):
        wb = openpyxl.load_workbook(file_path)
        # get the right sheet in the file (the second one)
        wb.active = 1
        sheet = wb.active
        columns = [
            sheet.cell(row=1, column=i).value for i in range(1, sheet.max_column + 1)
        ]
        idx_diagnosis_type = columns.index("Diagnosis Type")
        idx_diagnosis = columns.index("Diagnosis")
        idx_other_diagnosis = columns.index("Other Diagnosis Description")
        idx_subjectID = columns.index("CBTN Subject ID")
        idx_diagnosis_time = columns.index("Age at Diagnosis")
        idx_clinical_status = columns.index("Clinical Status")
        idx_overall_survival = columns.index("Overall Survival")
        idx_progression_free_survival = columns.index("Progression Free Survival")
        idx_gender = columns.index("Gender")
        # this points to the histopathology samples
        idx_CBTN_specimen_group = columns.index("CBTN Specimen Group ID")

        # get all the unique subject IDs, diagnostic types, diagnostic time, clinical_status, Overall survival, progression free survival
        # here collect all the data
        aus_subject_IDs = []
        aus_subject_diagnostic_types = []
        aus_subject_diagnostic_time = []
        aus_subject_diagnosis = []
        aus_subject_other_diagnosis = []
        aus_subject_clinical_status = []
        aus_overall_survival = []
        aus_progression_free_survival = []
        aus_gender = []
        aus_CBTN_specimen_group = []

        for row in range(2, sheet.max_row):
            if sheet.cell(row=row, column=idx_subjectID + 1).value != None:
                # get values
                aus_subject_IDs.append(
                    sheet.cell(row=row, column=idx_subjectID + 1).value
                )
                aus_subject_diagnostic_types.append(
                    sheet.cell(row=row, column=idx_diagnosis_type + 1).value
                )
                aus_subject_diagnostic_time.append(
                    sheet.cell(row=row, column=idx_diagnosis_time + 1).value
                )
                aus_subject_diagnosis.append(
                    sheet.cell(row=row, column=idx_diagnosis + 1).value
                )
                aus_subject_other_diagnosis.append(
                    sheet.cell(row=row, column=idx_other_diagnosis + 1).value
                )
                aus_subject_clinical_status.append(
                    sheet.cell(row=row, column=idx_clinical_status + 1).value
                )
                aus_overall_survival.append(
                    sheet.cell(row=row, column=idx_overall_survival + 1).value
                )
                aus_progression_free_survival.append(
                    sheet.cell(row=row, column=idx_progression_free_survival + 1).value
                )
                aus_gender.append(sheet.cell(row=row, column=idx_gender + 1).value)
                aus_CBTN_specimen_group.append(
                    sheet.cell(row=row, column=idx_CBTN_specimen_group + 1).value
                )

        # reorganize the information and save
        unique_subject_IDs = list(dict.fromkeys(aus_subject_IDs))
        clinical_info[tumor_type] = {}
        for s in unique_subject_IDs:
            # for this subject, get the diagnostic type and times, order them and save
            # get also a refined version of the tumor type not based on the
            # folder where the file is but on the disgnosis in the csv file

            # get the indexes belonging to this subject
            aus_subj_idx = [
                i for i, _ in enumerate(aus_subject_IDs) if aus_subject_IDs[i] == s
            ]

            # gather information
            subject_diagnostic_types = [
                aus_subject_diagnostic_types[i] for i in aus_subj_idx
            ]
            subject_diagnostic_time = [
                aus_subject_diagnostic_time[i] for i in aus_subj_idx
            ]
            subject_clinical_status = [
                aus_subject_clinical_status[i] for i in aus_subj_idx
            ]
            subject_diagnosis = [aus_subject_diagnosis[i] for i in aus_subj_idx]
            subject_other_diagnosis = [
                aus_subject_other_diagnosis[i] for i in aus_subj_idx
            ]
            subject_overall_survival = [aus_overall_survival[i] for i in aus_subj_idx]
            subject_free_survival = [
                aus_progression_free_survival[i] for i in aus_subj_idx
            ]
            subject_gender = aus_gender[aus_subj_idx[0]]

            subject_CBTN_specimen_group = [
                aus_CBTN_specimen_group[i] for i in aus_subj_idx
            ]

            """
            Note that in some cases the time is set as "Not reported". For these
            cases, the diagnosis is set to 0 if Clinical status is == "Alive" and
            if Diagnostic type contains "Initial". Else, the session is set to
            be the last one.
            """
            for i, t in enumerate(subject_diagnostic_time):
                if t == "Not Reported":
                    if all(
                        [
                            subject_clinical_status[i] == "Alive",
                            "Initial" in subject_diagnostic_types[i],
                        ]
                    ):
                        subject_diagnostic_time[i] = 0
                    else:
                        subject_diagnostic_time[i] = 10000000

            # order the two list based on the time
            (
                subject_diagnostic_time,
                subject_diagnostic_types,
                subject_overall_survival,
                subject_free_survival,
                subject_CBTN_specimen_group,
            ) = (
                list(t)
                for t in zip(
                    *sorted(
                        zip(
                            subject_diagnostic_time,
                            subject_diagnostic_types,
                            subject_overall_survival,
                            subject_free_survival,
                            subject_CBTN_specimen_group,
                        )
                    )
                )
            )

            # Write on the diagnostic time of fields having 0 or 10000000 that
            # this was automatically inserted
            # for i, t in enumerate(subject_diagnostic_time):
            #     if t == 0:
            #         subject_diagnostic_time[
            #             i
            #         ] = f"{subject_diagnostic_time[i+1]} (Set automatically)"
            #     elif t == 10000000:
            #         subject_diagnostic_time[i] = f"10000000 (Set automatically)"

            """
            Work on the diagnonis
            """
            diagnosis, description, grade = get_diagnosis(
                subject_diagnosis, subject_other_diagnosis
            )

            # save information
            clinical_info[tumor_type][s] = {
                "diagnostic_type": subject_diagnostic_types,
                "age_at_diagnosis": subject_diagnostic_time,
                "diagnosis_from_clinical_file": diagnosis,
                "diagnosis_description_from_clinical_file": description,
                "tumor_grade": grade,
                "overall_survival": subject_overall_survival,
                "progression_free_survival": subject_free_survival,
                "gender": subject_gender,
                "specimen_group": subject_CBTN_specimen_group,
            }
        wb.close()
    else:
        clinical_info[tumor_type] = None
